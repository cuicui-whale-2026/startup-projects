# -*- coding: utf-8 -*-
"""
D方向 产品3：计算机考公岗位分析表 Excel
数据来源：2026 国考公开信息（华图/本地宝/职位库等）
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\31125\Doubao\chats\2026-09-01\new-chat-1\startup-projects\计算机考公岗位分析表.xlsx"

wb = Workbook()

HEADER = "17365D"
ACCENT = "5B9BD5"
LIGHT = "DEEBF7"
GREEN = "C6EFCE"
ORANGE = "FFEB9C"
GRAY = "F2F2F2"

thin = Side(style="thin", color="B8C7D9")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def font(size=11, bold=False, color="000000"):
    return Font(name="微软雅黑", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center(cell, wrap=True):
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def make_header(ws, row, headers, color=ACCENT):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = font(11, True, "FFFFFF")
        c.fill = fill(color)
        center(c)
        c.border = border_all
    ws.row_dimensions[row].height = 26

def write_row(ws, row, values, fill_color=None):
    for col, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = font(10)
        c.border = border_all
        c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")
        if fill_color:
            c.fill = fill(fill_color)
    ws.row_dimensions[row].height = 30

# ========== Sheet1: 总览 ==========
ws = wb.active
ws.title = "岗位总览"
ws.merge_cells("A1:D1")
ws["A1"] = "2026 国考 · 计算机类专业岗位分析总览"
ws["A1"].font = font(16, True, "FFFFFF")
ws["A1"].fill = fill(HEADER)
center(ws["A1"])
ws.row_dimensions[1].height = 34

ws.merge_cells("A2:D2")
ws["A2"] = "数据来源：2026 国家公务员考试职位表公开信息（华图教育/本地宝/职位库）｜供选岗参考，以官方公告为准"
ws["A2"].font = font(9, False, "FFFFFF")
ws["A2"].fill = fill(ACCENT)
center(ws["A2"])
ws.row_dimensions[2].height = 22

rows = [
    ("可报岗位数", "约 4500+ 个", "计算机类（含软工/大数据/网安等）", ""),
    ("计划招录", "约 8000+ 人", "华图口径：岗位 4567 / 招录 10456", ""),
    ("本科岗位占比", "约 92%", "计算机可报岗位中本科占九成以上", ""),
    ("面向应届生", "近 70%", "应届生身份优势巨大", ""),
    ("主要系统", "税务/海关/金融监管/审计署/国家数据局/统计局", "岗位分化大，选岗决定上岸难度", ""),
]
headers = ["指标", "数据", "说明", "备注"]
make_header(ws, 4, headers)
r = 5
for row in rows:
    write_row(ws, r, list(row))
    r += 1

# 结论区
ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=4)
c = ws.cell(row=r+1, column=1, value="⭐ 核心结论：计算机是国考工科天花板，岗位多、应届优势大，但岗位分化严重——选岗选错直接影响上岸")
c.font = font(11, True, "17365D")
c.fill = fill(LIGHT)
c.alignment = Alignment(vertical="center", wrap_text=True)
ws.row_dimensions[r+1].height = 34

for col, w in zip("ABCD", [18, 22, 40, 12]):
    ws.column_dimensions[col].width = w

# ========== Sheet2: 可报岗位示例 ==========
ws2 = wb.create_sheet("可报岗位示例")
ws2.merge_cells("A1:F1")
ws2["A1"] = "2026 国考 · 计算机专业可报岗位示例（真实公开数据）"
ws2["A1"].font = font(14, True, "FFFFFF")
ws2["A1"].fill = fill(HEADER)
center(ws2["A1"])
ws2.row_dimensions[1].height = 30

headers2 = ["部门", "职位", "学历要求", "专业要求", "招录人数", "备注"]
make_header(ws2, 3, headers2)
jobs = [
    ("国家审计署", "审计业务处一级主任科员及以下", "硕士及以上", "计算机科学与技术/软件工程等", 2, "市(地)级"),
    ("公安部", "三处一级警长及以下", "硕士及以上", "0812计算机/0854电子信息/0839网安", 4, "中央级,报名91"),
    ("税务系统", "监管(电子信息化类)一级行政执法员", "仅限本科", "计算机/软工/网工/信息安全/人工智能", 4, "县(区)级,行政执法类"),
    ("海关系统", "科技管理一级行政执法员", "仅限本科", "计算机科学与技术/软工/网工/物联网等", 1, "辽阳海关,报名134"),
    ("统计系统", "国家统计局调查队一级科员", "仅限本科", "数学/统计学类(部分岗位)", 1, "北京丰台"),
    ("情报技术系统", "情报技术处一级警长及以下", "硕士及以上", "0812计算机/0835软工/0854电子信息", 5, "市(地)级,行政执法类"),
]
r = 4
for job in jobs:
    write_row(ws2, r, list(job))
    r += 1

for col, w in zip("ABCDEF", [16, 32, 14, 34, 10, 16]):
    ws2.column_dimensions[col].width = w

# ========== Sheet3: 报考策略 ==========
ws3 = wb.create_sheet("报考策略")
ws3.merge_cells("A1:C1")
ws3["A1"] = "计算机专业考公选岗策略"
ws3["A1"].font = font(14, True, "FFFFFF")
ws3["A1"].fill = fill(HEADER)
center(ws3["A1"])
ws3.row_dimensions[1].height = 30

strategy = [
    ("选岗原则", "优先应届可报 + 专业对口 + 招录人数多", "招3人以上的岗位进面概率更高"),
    ("学历卡位", "仅限本科的岗位竞争更小", "避开'硕士及以上'扎堆岗位"),
    ("系统选择", "税务招录量大但报名人数也多", "数据局/网信办等新兴系统竞争相对小"),
    ("地域策略", "发达地区(北上广深)分数普遍更高", "可考虑家乡或二线城市曲线上岸"),
    ("行政执法类", "比综合管理类多考一门专业课? 非也", "行政执法类申论侧重执法场景,需针对性准备"),
    ("时间规划", "应届身份只有一次,务必珍惜", "国考11月底笔试,提前6个月系统备考"),
]
headers3 = ["维度", "建议", "说明"]
make_header(ws3, 3, headers3)
r = 4
for s in strategy:
    write_row(ws3, r, list(s))
    r += 1

for col, w in zip("ABC", [16, 40, 36]):
    ws3.column_dimensions[col].width = w

# ========== Sheet4: 备考计划 ==========
ws4 = wb.create_sheet("备考节奏")
ws4.merge_cells("A1:C1")
ws4["A1"] = "考公备考节奏参考（6个月）"
ws4["A1"].font = font(14, True, "FFFFFF")
ws4["A1"].fill = fill(HEADER)
center(ws4["A1"])
ws4.row_dimensions[1].height = 30

plan = [
    ("第1-2月", "打基础", "行测五大模块过一遍，申论建立框架"),
    ("第3-4月", "强化刷题", "行测每天2h+专项突破，申论每周3篇"),
    ("第5月", "真题模考", "近5年真题限时模考，错题复盘"),
    ("第6月", "冲刺", "全真模拟+时政热点+申论热点背诵"),
]
headers4 = ["阶段", "重点", "具体动作"]
make_header(ws4, 3, headers4)
r = 4
for p in plan:
    write_row(ws4, r, list(p))
    r += 1

for col, w in zip("ABC", [16, 18, 50]):
    ws4.column_dimensions[col].width = w

ws.sheet_view.showGridLines = False
ws2.sheet_view.showGridLines = False
ws3.sheet_view.showGridLines = False
ws4.sheet_view.showGridLines = False

wb.save(OUT)
print(f"✅ 已生成: {OUT}")
print(f"Sheet 数: {len(wb.sheetnames)} -> {wb.sheetnames}")
