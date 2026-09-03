# -*- coding: utf-8 -*-
"""
生成「备考周计划表」Excel 模板（D方向可售产品样张）
- 周目标 + 每日任务
- 三色进度标记
- 完成率自动统计（公式）
- 复盘区
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

OUT = r"C:\Users\31125\Doubao\chats\2026-09-01\new-chat-1\startup-projects\D_备考周计划表.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "备考周计划"

# ---------- 配色 ----------
HEADER = "17365D"      # 深蓝表头
ACCENT = "5B9BD5"      # 主蓝
LIGHT = "DEEBF7"       # 浅蓝
TASK = "FFF2CC"        # 任务区浅黄
DONE = "C6EFCE"        # 完成绿
DOING = "FFEB9C"       # 进行中黄
TODO = "FFFFFF"        # 未开始白
GRAY = "F2F2F2"

thin = Side(style="thin", color="B8C7D9")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def font(size=11, bold=False, color="000000", name="微软雅黑"):
    return Font(name=name, size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center(cell, wrap=True):
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)

# ---------- 标题行 ----------
ws.merge_cells("A1:F1")
ws["A1"] = "备考周计划表"
ws["A1"].font = font(16, True, "FFFFFF")
ws["A1"].fill = fill(HEADER)
center(ws["A1"])
ws.row_dimensions[1].height = 32

# 副标题
ws.merge_cells("A2:F2")
ws["A2"] = "周目标：__________________　|　本周关键词：__________________"
ws["A2"].font = font(10, False, "FFFFFF")
ws["A2"].fill = fill(ACCENT)
center(ws["A2"])
ws.row_dimensions[2].height = 24

# ---------- 表头 ----------
headers = ["日期", "科目/任务", "具体内容", "完成度", "状态", "备注"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=3, column=col, value=h)
    c.font = font(11, True, "FFFFFF")
    c.fill = fill(ACCENT)
    center(c)
    c.border = border_all
ws.row_dimensions[3].height = 26

# ---------- 数据区（7天 × 每日3行） ----------
days = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
row = 4
task_rows = {}  # 记录每天任务行，用于完成率公式
day_start = {}
for d in days:
    day_start[d] = row
    # 日期合并列
    ws.merge_cells(start_row=row, start_column=1, end_row=row+2, end_column=1)
    dc = ws.cell(row=row, column=1, value=d)
    dc.font = font(11, True)
    dc.fill = fill(LIGHT)
    center(dc)
    dc.border = border_all
    # 补下方合并单元格边框
    for r in range(row, row+3):
        ws.cell(row=r, column=1).border = border_all

    for i in range(3):
        # 任务行
        t = ws.cell(row=row+i, column=2, value=f"任务 {i+1}")
        t.font = font(10)
        t.alignment = Alignment(vertical="center")
        t.border = border_all
        t.fill = fill(TASK)
        # 具体内容
        content = ws.cell(row=row+i, column=3, value="")
        content.border = border_all
        content.alignment = Alignment(vertical="center", wrap_text=True)
        # 完成度（百分比，默认空）
        pct = ws.cell(row=row+i, column=4, value=None)
        pct.number_format = "0%"
        pct.border = border_all
        center(pct)
        # 状态（下拉候选：未开始/进行中/完成）
        status = ws.cell(row=row+i, column=5, value="未开始")
        status.font = font(10)
        status.border = border_all
        center(status)
        # 备注
        note = ws.cell(row=row+i, column=6, value="")
        note.border = border_all
        note.alignment = Alignment(vertical="center", wrap_text=True)

    # 分隔线强化（每天结束行下边框加粗）
    for col in range(1, 7):
        ws.cell(row=row+2, column=col).border = Border(
            left=thin, right=thin, bottom=Side(style="medium", color="5B9BD5"))

    # 每天的行高
    for r in range(row, row+3):
        ws.row_dimensions[r].height = 28
    row += 3

# ---------- 汇总区 ----------
sum_row = row + 1
ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=3)
sc = ws.cell(row=sum_row, column=1, value="本周总完成率（自动统计）")
sc.font = font(12, True, "FFFFFF")
sc.fill = fill(HEADER)
center(sc)
ws.merge_cells(start_row=sum_row, start_column=4, end_row=sum_row, end_column=6)
total = ws.cell(row=sum_row, column=4)
# 完成率 = 完成状态数 / 总任务数
total.value = f'=COUNTIF(E4:E{row-1},"完成")/COUNTIF(B4:B{row-1},"任务*")'
total.number_format = "0%"
total.font = font(12, True, "FFFFFF")
total.fill = fill(HEADER)
center(total)
ws.row_dimensions[sum_row].height = 26

# ---------- 复盘区 ----------
review_row = sum_row + 2
ws.merge_cells(start_row=review_row, start_column=1, end_row=review_row, end_column=6)
rc = ws.cell(row=review_row, column=1, value="📝 本周复盘")
rc.font = font(12, True, "FFFFFF")
rc.fill = fill(ACCENT)
center(rc)
ws.row_dimensions[review_row].height = 26
for i in range(1, 5):
    rr = review_row + i
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=6)
    item = ["① 本周完成最好的事：", "② 本周最拖延/没做好的事：", "③ 下周改进计划：", "④ 一句话总结："]
    c = ws.cell(row=rr, column=1, value=item[i-1])
    c.font = font(10)
    c.fill = fill(GRAY)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[rr].height = 24

# ---------- 列宽 ----------
widths = {"A": 10, "B": 14, "C": 28, "D": 10, "E": 10, "F": 18}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# ---------- 条件格式：状态列着色（未开始白/进行中黄/完成绿） ----------
from openpyxl.formatting.rule import FormulaRule
rng = f"E4:E{row-1}"
ws.conditional_formatting.add(rng, FormulaRule(formula=['$E4="完成"'], fill=fill(DONE)))
ws.conditional_formatting.add(rng, FormulaRule(formula=['$E4="进行中"'], fill=fill(DOING)))

# 隐藏网格线，更干净
ws.sheet_view.showGridLines = False

wb.save(OUT)
print(f"已生成: {OUT}")
print(f"数据行: 4 到 {row-1}，共 {(row-4)//3} 天 × 3 任务 = {row-4} 行任务")
print(f"汇总行: {sum_row}")
